import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference

# Reads the Excel input file
input_file = r"C:\Users\maria\Desktop\ΠΡΟΤΖΕΚΤ1.ΤΟΨΙΣ\topsis_input.xlsx"
data_alternatives = pd.read_excel(input_file, sheet_name="Decision_Matrix", index_col=0)
data_weights = pd.read_excel(input_file, sheet_name="Criteria_Weights")
data_criteria_type = pd.read_excel(input_file, sheet_name="Criteria_Types")

# Retrieve data
criteria_labels = data_alternatives.columns.tolist()
alternative_labels = data_alternatives.index.tolist()
alternative_matrix = data_alternatives.values
criteria_importance = data_weights["Weight"].values
criteria_nature = data_criteria_type["Type"].apply(lambda x: 1 if x == "Benefit" else 0).values

# Decision matrix normalization
norm_matrix = alternative_matrix / np.sqrt((alternative_matrix**2).sum(axis=0))
norm_data = pd.DataFrame(norm_matrix, index=alternative_labels, columns=criteria_labels)

# Weighted normalized matrix
weighted_matrix = norm_matrix * criteria_importance
weighted_data = pd.DataFrame(weighted_matrix, index=alternative_labels, columns=criteria_labels)

# Calculate ideal solutions
best_solutions = [max(weighted_matrix[:, j]) if criteria_nature[j] == 1 else min(weighted_matrix[:, j]) for j in range(len(criteria_labels))]
worst_solutions = [min(weighted_matrix[:, j]) if criteria_nature[j] == 1 else max(weighted_matrix[:, j]) for j in range(len(criteria_labels))]

# Distance from ideal solutions
distance_best = np.sqrt(((weighted_matrix - best_solutions)**2).sum(axis=1))
distance_worst = np.sqrt(((weighted_matrix - worst_solutions)**2).sum(axis=1))

# Calculate relative closeness
closeness_scores = distance_worst / (distance_best + distance_worst)

# Ranking
sorted_indices = np.argsort(-closeness_scores)
ranking_positions = np.zeros_like(sorted_indices)
ranking_positions[sorted_indices] = np.arange(1, len(closeness_scores) + 1)

# Save results to DataFrame
ranking_results = pd.DataFrame({
    "Alternative": np.array(alternative_labels)[sorted_indices],
    "Relative Closeness": closeness_scores[sorted_indices],
    "Rank": np.arange(1, len(closeness_scores) + 1)
})

# Save to Excel
output_file = "topsis_final_results.xlsx"
with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
     ranking_results.to_excel(writer, sheet_name="Ranking", index=False)
     norm_data.to_excel(writer, sheet_name="Normalized_Matrix", index=True)
     weighted_data.to_excel(writer, sheet_name="Weighted_Matrix", index=True)

# Add charts to Excel
workbook = load_workbook(output_file)
chart_page = workbook.create_sheet(title="Charts")

# Relative closeness chart
chart = BarChart()
chart.title = "Relative Closeness of Alternatives"
chart.x_axis.title = "Alternatives"
chart.y_axis.title = "Relative Closeness"
data_values = Reference(workbook["Ranking"], min_col=2, max_col=2, min_row=1, max_row=len(alternative_labels) + 1)
data_categories = Reference(workbook["Ranking"], min_col=1, max_col=1, min_row=2, max_row=len(alternative_labels) + 1)
chart.add_data(data_values, titles_from_data=True)
chart.set_categories(data_categories)
chart_page.add_chart(chart, "A1")

workbook.save(output_file)

# Python visualization
plt.figure(figsize=(10, 6))
plt.bar(ranking_results["Alternative"], ranking_results["Relative Closeness"], color="lightblue", edgecolor="black")
plt.title("Relative Closeness of Alternatives")
plt.xlabel("Alternatives")
plt.ylabel("Relative Closeness")
plt.xticks(rotation=45)
plt.grid(axis='y', linestyle='--', alpha=0.7)
plt.show()

print(f"Results saved to '{output_file}'.")
